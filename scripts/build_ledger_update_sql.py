"""
把审核后的台账工作簿转换成 SQL 更新语句。

适用输入：
- build_ledger_review_workbook.py 生成并编辑后的 .xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from import_historical_data import escape_sql


YES_VALUES = {'是', 'yes', 'y', '1', 'true', 'TRUE'}


def parse_args():
    parser = argparse.ArgumentParser(description='把审核工作簿转换成 SQL 更新语句')
    parser.add_argument('--input-xlsx', required=True, help='审核后的工作簿')
    parser.add_argument('--output', help='输出 SQL 路径')
    return parser.parse_args()


def normalize(value):
    if value is None:
        return ''
    return str(value).strip()


def to_decimal(value):
    if value is None or value == '':
        return None
    return float(value)


def is_yes(value) -> bool:
    return normalize(value) in YES_VALUES


def build_update(table_name: str, record_id: str, changes: dict[str, object]):
    if not changes:
        return None

    parts = []
    for key, value in changes.items():
        if isinstance(value, (int, float)):
            parts.append(f"{key} = {value:.2f}")
        else:
            parts.append(f"{key} = '{escape_sql(value)}'")
    return f"UPDATE {table_name} SET {', '.join(parts)} WHERE id = '{escape_sql(record_id)}';"


def main():
    args = parse_args()
    input_xlsx = Path(args.input_xlsx)
    if not input_xlsx.exists():
        raise FileNotFoundError(f'工作簿不存在: {input_xlsx}')

    wb = load_workbook(input_xlsx, data_only=True)
    if '台账审核' not in wb.sheetnames:
        raise ValueError('工作簿中缺少“台账审核”工作表')

    ws = wb['台账审核']
    sql_lines = [
        '-- LuckCup 台账审核回写 SQL',
        f'-- source: {input_xlsx.name}',
        '',
    ]
    update_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        month, date, table_name, entry_mode, direction, current_name, original_name, current_amount, current_note, is_auto, record_id, created_at, target_name, target_amount, target_note, should_update, review_note = row[:17]

        if not record_id or not is_yes(should_update):
            continue

        current_name = normalize(current_name)
        target_name = normalize(target_name) or current_name
        current_note = normalize(current_note)
        target_note = normalize(target_note) if target_note is not None else current_note
        current_amount = to_decimal(current_amount)
        target_amount = to_decimal(target_amount)
        if target_amount is None:
            target_amount = current_amount

        changes = {}

        if table_name == 'daily_income':
            if target_name != current_name:
                changes['platform'] = target_name
            if target_amount != current_amount:
                changes['amount'] = target_amount
        elif table_name == 'expenses':
            if target_name != current_name:
                changes['category'] = target_name
            if target_amount != current_amount:
                changes['amount'] = target_amount
            if target_note != current_note:
                changes['note'] = target_note
        elif table_name == 'backend_entries':
            if target_name != current_name:
                changes['category'] = target_name
            if target_amount != current_amount:
                changes['amount'] = target_amount
            if target_note != current_note:
                changes['note'] = target_note
        else:
            continue

        statement = build_update(table_name, normalize(record_id), changes)
        if statement:
            sql_lines.append(statement)
            update_count += 1

    output_path = Path(args.output) if args.output else input_xlsx.with_name(f'{input_xlsx.stem}_回写.sql')
    output_path.write_text('\n'.join(sql_lines), encoding='utf-8')
    print(f'✅ 回写 SQL 已生成: {output_path}')
    print(f'   共 {update_count} 条 UPDATE')


if __name__ == '__main__':
    main()
