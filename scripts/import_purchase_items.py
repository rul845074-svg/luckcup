"""
LuckCup 报货货品导入脚本
从毛庄报货 xlsx 文件提取货品名称，生成 SQL 导入文件。

用法: python3 scripts/import_purchase_items.py
输出: scripts/output/purchase_items_import.sql
"""

import uuid
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")
from openpyxl import load_workbook

BASE = Path("/Users/zhanghu/Library/Mobile Documents/com~apple~Numbers/Documents/毛庄清单")
OUTPUT_DIR = Path(__file__).parent / "output"

# 非货品行（统计行/备注行）
NON_GOODS = {
    "记", "废", "营业额", "订货额", "周边", "进销比",
    "订单数", "客单价", "坪效", "人员", "实收",
}

# 周边货物
PERIPHERAL_GOODS = {
    "挂耳", "小圆饼干", "碱水面包干", "咖啡豆小饼干",
    "原味拿铁", "椰椰拿铁", "玉米球",
}

# 包材耗材
PACKAGE_GOODS = {
    "粗吸管", "细吸管", "500塑杯", "PET高杯", "98平盖",
    "16A纸杯", "12A纸杯", "90注塑盖", "球型盖", "90平盖",
    "防漏纸", "单杯塑料袋", "单杯纸袋", "双杯纸袋", "四杯塑袋",
    "双杯保温袋", "四杯保温袋", "双杯纸托", "双杯塑托", "四杯塑托",
    "U型杯", "无纺布", "小票纸", "标签纸", "圣代勺",
}


def classify(name):
    if name in NON_GOODS:
        return None  # 跳过
    if name in PERIPHERAL_GOODS:
        return "周边货物"
    if name in PACKAGE_GOODS:
        return "包材耗材"
    return "原料货品"


def escape_sql(s):
    return str(s).replace("\\", "\\\\").replace("'", "\\'")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    sql_lines = []
    sql_lines.append("-- LuckCup 报货货品导入")
    sql_lines.append("-- 自动生成")
    sql_lines.append("")
    sql_lines.append("SET @shop_id = (SELECT id FROM shops LIMIT 1);")
    sql_lines.append("")

    seen = set()

    for xlsx_name, year in [("毛庄报货2025.xlsx", "2025"), ("毛庄报货2026.xlsx", "2026")]:
        filepath = BASE / xlsx_name
        if not filepath.exists():
            print(f"跳过: {xlsx_name} 不存在")
            continue

        wb = load_workbook(filepath, data_only=True)
        print(f"读取 {xlsx_name}, 工作表: {wb.sheetnames}")

        for ws in wb.worksheets:
            # 尝试从工作表名提取月份
            month_name = ws.title.strip()
            # 推断 source_month
            import re
            m = re.search(r"(\d+)", month_name)
            source_month = f"{year}-{int(m.group(1)):02d}" if m else f"{year}-01"

            for row in range(2, ws.max_row + 1):
                val = ws.cell(row, 1).value
                if val is None:
                    continue
                name = str(val).strip()
                if not name:
                    continue

                cat = classify(name)
                if cat is None:
                    continue  # 统计行

                if name in seen:
                    continue
                seen.add(name)

                uid = str(uuid.uuid4())
                sql_lines.append(
                    f"INSERT IGNORE INTO purchase_items (id, shop_id, name, category, source_month) "
                    f"VALUES ('{uid}', @shop_id, '{escape_sql(name)}', '{escape_sql(cat)}', '{source_month}');"
                )

    output_path = OUTPUT_DIR / "purchase_items_import.sql"
    output_path.write_text("\n".join(sql_lines), encoding="utf-8")
    print(f"\n✅ SQL 已生成: {output_path}")
    print(f"   共 {len(seen)} 个货品")


if __name__ == "__main__":
    main()
