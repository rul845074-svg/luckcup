"""
LuckCup 单月补导脚本（XLSX 版）

适用场景：
- 直接从整理好的 Excel 工作簿中读取某个月的数据
- 避免 .numbers 解析异常

用法示例：
python3 scripts/import_month_from_xlsx.py \
  --month 2026-02 \
  --xlsx "/path/to/核对清单2026-2月.xlsx" \
  --shop-id "xxx-shop-id"
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from import_month_from_csv import build_sql, is_summary_label, normalize, parse_amount, parse_date


OUTPUT_DIR = Path(__file__).parent / "output"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从 Excel 工作簿生成单月补导 SQL")
    parser.add_argument("--month", required=True, help="月份，格式 YYYY-MM，例如 2026-02")
    parser.add_argument("--xlsx", required=True, help="Excel 文件路径")
    parser.add_argument("--shop-id", help="指定导入到哪个 shop_id，强烈建议传这个参数")
    parser.add_argument(
        "--output",
        help="输出 SQL 路径，默认 scripts/output/month_import_YYYY_MM.sql",
    )
    return parser.parse_args()


def parse_sheet(ws, year: int, month: int, entry_type: str):
    current_date = None
    entries = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 4:
            continue

        date_raw = row[1]
        amount_raw = row[2]
        category_raw = row[3]

        category = normalize(category_raw)
        if not category or category in {"类别", "合计", "总计"}:
            continue
        if is_summary_label(category):
            continue

        amount = parse_amount(amount_raw)
        if amount is None:
            continue

        parsed_date = parse_date(date_raw, year, month)
        if parsed_date is not None:
            current_date = parsed_date
        if current_date is None:
            current_date = f"{year:04d}-{month:02d}-01"

        entries.append(
            {
                "date": current_date,
                "amount": amount,
                "original_category": category,
                "type": entry_type,
            }
        )

    return entries


def main():
    args = parse_args()
    year, month = map(int, args.month.split("-"))
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)
    income_sheet_name = f"{month}月收入"
    expense_sheet_name = f"{month}月支出"

    if income_sheet_name not in wb.sheetnames:
        raise ValueError(f"找不到工作表: {income_sheet_name}")
    if expense_sheet_name not in wb.sheetnames:
        raise ValueError(f"找不到工作表: {expense_sheet_name}")

    entries = []
    entries.extend(parse_sheet(wb[income_sheet_name], year, month, "income"))
    entries.extend(parse_sheet(wb[expense_sheet_name], year, month, "expense"))

    sql_lines, stats = build_sql(args.month, entries, args.shop_id)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = Path(args.output) if args.output else OUTPUT_DIR / f"month_import_{year}_{month:02d}.sql"
    output_path.write_text("\n".join(sql_lines), encoding="utf-8")

    print(f"✅ SQL 已生成: {output_path}")
    print(f"   前台收入: {stats['income_rows']} 条")
    print(f"   前台支出: {stats['expense_rows']} 条")
    print(f"   后台补录: {stats['backend_rows']} 条")
    if stats["unmapped"]:
        print("   未映射条目:")
        for name, count in stats["unmapped"].most_common():
            print(f"   - {name}: {count}")


if __name__ == "__main__":
    main()
