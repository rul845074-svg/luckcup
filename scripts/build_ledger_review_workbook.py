"""
把系统导出的月度台账 CSV 转成可审核、可下拉选择的 Excel 工作簿。

推荐流程：
1. 在系统“盈亏分析”页点击“导出本月台账 CSV”
2. 运行本脚本生成审核工作簿
3. 在 Excel / Numbers 中修改“目标系统名称 / 目标金额 / 目标备注 / 是否更新”
4. 再用 build_ledger_update_sql.py 生成回写 SQL
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from import_historical_data import EXPENSE_MAP, INCOME_MAP


DEFAULT_PLATFORMS = ['美团团购', '美团外卖', '外卖其他', '抖音团购', '小程序', '门店收银']
DEFAULT_EXPENSE_CATEGORIES = ['原料货品', '周边货物', '工资', '房租', '水电', '其他支出', '活动', '物业', '其他']

HEADERS = [
    '月份',
    '日期',
    '来源表',
    '录入方式',
    '收支类型',
    '当前系统名称',
    '原始名称',
    '当前金额',
    '当前备注',
    '是否自动',
    '记录ID',
    '创建时间',
    '目标系统名称',
    '目标金额',
    '目标备注',
    '是否更新',
    '审核备注',
]


def parse_args():
    parser = argparse.ArgumentParser(description='生成月度台账审核工作簿')
    parser.add_argument('--input-csv', required=True, help='系统导出的月度台账 CSV')
    parser.add_argument('--output', help='输出 xlsx 路径')
    return parser.parse_args()


def unique_in_order(values):
    seen = set()
    result = []
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def load_rows(path: Path):
    with path.open('r', encoding='utf-8-sig', newline='') as f:
        reader = csv.DictReader(f)
        return list(reader)


def build_option_lists(rows):
    platforms = unique_in_order(DEFAULT_PLATFORMS + [
        row['系统类目'] for row in rows if row['来源表'] == 'daily_income'
    ])
    expense_categories = unique_in_order(DEFAULT_EXPENSE_CATEGORIES + [
        row['系统类目'] for row in rows if row['来源表'] == 'expenses'
    ])

    backend_income_categories = unique_in_order([
        target for original, (mode, target) in INCOME_MAP.items() if mode == '后台'
    ] + [
        row['系统类目']
        for row in rows
        if row['来源表'] == 'backend_entries' and row['收支类型'] == '收入'
    ])

    backend_expense_categories = unique_in_order([
        target for original, (mode, target) in EXPENSE_MAP.items() if mode == '后台'
    ] + [
        row['系统类目']
        for row in rows
        if row['来源表'] == 'backend_entries' and row['收支类型'] == '支出'
    ])

    return {
        'platforms': platforms,
        'expense_categories': expense_categories,
        'backend_income_categories': backend_income_categories,
        'backend_expense_categories': backend_expense_categories,
        'yes_no': ['否', '是'],
    }


def fill_instruction_sheet(ws):
    lines = [
        '1. 不要改“记录ID / 来源表 / 当前系统名称”等灰底列。',
        '2. 需要调整时，只改“目标系统名称 / 目标金额 / 目标备注”，并把“是否更新”改成“是”。',
        '3. 目标系统名称可以用下拉选择，避免手打出错。',
        '4. 改完后运行 build_ledger_update_sql.py，把工作簿转成 SQL 再执行。',
        '5. daily_income 改平台或日期时，若撞上同日同平台唯一键，SQL 会报重复，需要人工合并。',
    ]
    ws['A1'] = 'LuckCup 台账审核说明'
    ws['A1'].font = Font(bold=True, size=14)
    for index, line in enumerate(lines, start=3):
      ws.cell(index, 1).value = line
    ws.column_dimensions['A'].width = 80


def fill_options_sheet(ws, options):
    columns = {
        'A': ('前台收入平台', options['platforms']),
        'B': ('前台支出类目', options['expense_categories']),
        'C': ('后台收入类目', options['backend_income_categories']),
        'D': ('后台支出类目', options['backend_expense_categories']),
        'E': ('是否更新', options['yes_no']),
    }
    for col, (title, values) in columns.items():
        ws[f'{col}1'] = title
        ws[f'{col}1'].font = Font(bold=True)
        for row_index, value in enumerate(values, start=2):
            ws[f'{col}{row_index}'] = value
        ws.column_dimensions[col].width = 24


def apply_header_style(ws):
    fill = PatternFill('solid', fgColor='F4E7DA')
    locked_fill = PatternFill('solid', fgColor='F3F4F6')
    for col_index, title in enumerate(HEADERS, start=1):
        cell = ws.cell(1, col_index)
        cell.value = title
        cell.font = Font(bold=True)
        cell.fill = fill if col_index >= 13 else locked_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


def add_row_validations(ws, row_index, row, option_lengths):
    option_sheet = '选项'

    if row['来源表'] == 'daily_income':
        formula = f'={option_sheet}!$A$2:$A${option_lengths["platforms"] + 1}'
    elif row['来源表'] == 'expenses':
        formula = f'={option_sheet}!$B$2:$B${option_lengths["expense_categories"] + 1}'
    elif row['收支类型'] == '收入':
        formula = f'={option_sheet}!$C$2:$C${option_lengths["backend_income_categories"] + 1}'
    else:
        formula = f'={option_sheet}!$D$2:$D${option_lengths["backend_expense_categories"] + 1}'

    dv_name = DataValidation(type='list', formula1=formula, allow_blank=True)
    ws.add_data_validation(dv_name)
    dv_name.add(f'M{row_index}')

    dv_yes_no = DataValidation(type='list', formula1=f'={option_sheet}!$E$2:$E$3', allow_blank=False)
    ws.add_data_validation(dv_yes_no)
    dv_yes_no.add(f'P{row_index}')


def main():
    args = parse_args()
    input_csv = Path(args.input_csv)
    if not input_csv.exists():
        raise FileNotFoundError(f'CSV 不存在: {input_csv}')

    rows = load_rows(input_csv)
    if not rows:
        raise ValueError('CSV 没有可用数据')

    options = build_option_lists(rows)
    option_lengths = {key: len(value) for key, value in options.items()}

    wb = Workbook()
    ws_review = wb.active
    ws_review.title = '台账审核'
    ws_help = wb.create_sheet('说明')
    ws_options = wb.create_sheet('选项')

    fill_instruction_sheet(ws_help)
    fill_options_sheet(ws_options, options)
    apply_header_style(ws_review)

    for row_index, row in enumerate(rows, start=2):
        review_values = [
            row.get('月份', ''),
            row.get('日期', ''),
            row.get('来源表', ''),
            row.get('录入方式', ''),
            row.get('收支类型', ''),
            row.get('系统类目', ''),
            row.get('原始名称', ''),
            row.get('金额', ''),
            row.get('备注', ''),
            row.get('是否自动', ''),
            row.get('记录ID', ''),
            row.get('创建时间', ''),
            row.get('系统类目', ''),
            row.get('金额', ''),
            row.get('备注', ''),
            '否',
            '',
        ]

        for col_index, value in enumerate(review_values, start=1):
            ws_review.cell(row_index, col_index).value = value

        add_row_validations(ws_review, row_index, row, option_lengths)

    ws_review.freeze_panes = 'A2'
    ws_review.auto_filter.ref = ws_review.dimensions
    widths = {
        'A': 10, 'B': 12, 'C': 16, 'D': 12, 'E': 10, 'F': 18, 'G': 18,
        'H': 12, 'I': 22, 'J': 10, 'K': 38, 'L': 20, 'M': 18, 'N': 12,
        'O': 22, 'P': 10, 'Q': 24,
    }
    for col, width in widths.items():
        ws_review.column_dimensions[col].width = width

    output_path = Path(args.output) if args.output else input_csv.with_name(f'{input_csv.stem}_审核工作簿.xlsx')
    wb.save(output_path)
    print(f'✅ 审核工作簿已生成: {output_path}')


if __name__ == '__main__':
    main()
